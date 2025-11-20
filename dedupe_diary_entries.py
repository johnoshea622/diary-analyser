#!/usr/bin/env python3
"""
Parse every diary workbook/sheet, collect personnel and production activities,
and emit CSV reports that highlight duplicates vs single-instance entries per day.

Example:
    python dedupe_diary_entries.py --root "." --output-dir analysis
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Set, Tuple

from openpyxl import load_workbook

import build_diary_database as diary


@dataclass(frozen=True)
class ActivityEntry:
    diary_date: date
    activity: str
    source_file: str
    worksheet: str

    @property
    def source_label(self) -> str:
        return f"{self.source_file}::{self.worksheet}"


@dataclass(frozen=True)
class PersonnelEntry:
    diary_date: date
    team: str
    name: str
    position: str
    hours: float
    source_file: str
    worksheet: str

    @property
    def source_label(self) -> str:
        return f"{self.source_file}::{self.worksheet}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract diary entries and flag duplicates.")
    parser.add_argument("--root", default=".", help="Root folder that contains the Excel workbooks.")
    parser.add_argument(
        "--output-dir",
        default="analysis",
        help="Directory where CSV reports will be written (created if needed).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    activities, personnel, date_sources, errors = gather_entries(root)

    activities_rows, activities_unique = annotate_activity_entries(activities, date_sources)
    personnel_rows, personnel_unique = annotate_personnel_entries(personnel, date_sources)

    write_csv(
        output_dir / "activities_entries.csv",
        [
            "diary_date",
            "activity_text",
            "source_file",
            "worksheet",
            "unique_to_source",
            "status",
            "all_sources_for_date",
        ],
        activities_rows,
    )
    write_csv(
        output_dir / "activities_unique.csv",
        [
            "diary_date",
            "activity_text",
            "sources_present",
            "sources_missing",
            "status",
            "report_copies_for_date",
        ],
        activities_unique,
    )
    write_csv(
        output_dir / "personnel_entries.csv",
        [
            "diary_date",
            "team",
            "name",
            "position",
            "hours",
            "source_file",
            "worksheet",
            "unique_to_source",
            "status",
            "all_sources_for_date",
        ],
        personnel_rows,
    )
    write_csv(
        output_dir / "personnel_unique.csv",
        [
            "diary_date",
            "team",
            "name",
            "position",
            "hours",
            "sources_present",
            "sources_missing",
            "status",
            "report_copies_for_date",
        ],
        personnel_unique,
    )

    print("Processed diary entries.")
    print(f"  Activities captured: {len(activities)}")
    print(f"  Personnel captured:  {len(personnel)}")
    if errors:
        print("Issues encountered:")
        for item in errors:
            print(f"  - {item}")
    else:
        print("  No parsing issues detected.")
    print(f"CSV output directory: {output_dir}")


def gather_entries(root: Path) -> Tuple[List[ActivityEntry], List[PersonnelEntry], Dict[date, Set[str]], List[str]]:
    activities: List[ActivityEntry] = []
    personnel: List[PersonnelEntry] = []
    date_sources: Dict[date, Set[str]] = defaultdict(set)
    errors: List[str] = []

    files = diary.iter_excel_files(root)
    if not files:
        errors.append(f"No Excel files found under {root}")
        return activities, personnel, date_sources, errors

    for file_path in files:
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - defensive logging
            errors.append(f"Failed to open {file_path}: {exc}")
            continue
        relative_file = _relative_to_root(file_path, root)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(diary.iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = diary.extract_diary_date(rows)
            if diary_date is None:
                errors.append(f"Skipping {relative_file}::{sheet_name} (no diary date found)")
                continue
            source_label = f"{relative_file}::{sheet_name}"
            date_sources[diary_date].add(source_label)
            for activity_text in extract_activity_cells(rows):
                activities.append(
                    ActivityEntry(
                        diary_date=diary_date,
                        activity=activity_text,
                        source_file=relative_file,
                        worksheet=sheet_name,
                    )
                )
            for team, name, position, hours in diary.extract_personnel(rows):
                personnel.append(
                    PersonnelEntry(
                        diary_date=diary_date,
                        team=team.strip(),
                        name=name.strip(),
                        position=(position or "").strip(),
                        hours=float(hours or 0.0),
                        source_file=relative_file,
                        worksheet=sheet_name,
                    )
                )
    return activities, personnel, date_sources, errors


def extract_activity_cells(rows: Sequence[diary.SheetRow]) -> List[str]:
    activities: List[str] = []
    in_section = False
    for row in rows:
        row_upper = row.upper
        if not in_section:
            if "PRODUCTION" in row_upper:
                in_section = True
            continue
        if "PHOTOS" in row_upper:
            break
        if "COMMUNICATIONS" in row_upper:
            continue
        for cell in row.text:
            if not cell:
                continue
            if "PRODUCTION" in cell.upper():
                continue
            for chunk in split_multiline(cell):
                cleaned = chunk.strip()
                if cleaned and cleaned.upper() not in {"PHOTOS"}:
                    activities.append(cleaned)
    return activities


def annotate_activity_entries(
    entries: Sequence[ActivityEntry], date_sources: Dict[date, Set[str]]
) -> Tuple[List[dict], List[dict]]:
    grouped: Dict[Tuple[date, str], List[ActivityEntry]] = defaultdict(list)
    for entry in entries:
        grouped[(entry.diary_date, normalize_text(entry.activity))].append(entry)
    rows: List[dict] = []
    summary: List[dict] = []
    for (entry_date, _), items in grouped.items():
        canonical = items[0].activity
        all_sources = sorted(date_sources.get(entry_date, set()))
        sources_present = sorted({item.source_label for item in items})
        missing = sorted(set(all_sources) - set(sources_present))
        status, unique_flag = describe_presence(sources_present, missing, len(all_sources))
        for item in items:
            rows.append(
                {
                    "diary_date": entry_date.isoformat(),
                    "activity_text": item.activity,
                    "source_file": item.source_file,
                    "worksheet": item.worksheet,
                    "unique_to_source": unique_flag,
                    "status": status,
                    "all_sources_for_date": "; ".join(all_sources),
                }
            )
        summary.append(
            {
                "diary_date": entry_date.isoformat(),
                "activity_text": canonical,
                "sources_present": "; ".join(sources_present),
                "sources_missing": "; ".join(missing),
                "status": status,
                "report_copies_for_date": len(all_sources),
            }
        )
    rows.sort(key=lambda row: (row["diary_date"], row["activity_text"], row["source_file"]))
    summary.sort(key=lambda row: (row["diary_date"], row["activity_text"]))
    return rows, summary


def annotate_personnel_entries(
    entries: Sequence[PersonnelEntry], date_sources: Dict[date, Set[str]]
) -> Tuple[List[dict], List[dict]]:
    grouped: Dict[Tuple[date, str, str, str, float], List[PersonnelEntry]] = defaultdict(list)
    for entry in entries:
        grouped[
            (
                entry.diary_date,
                normalize_text(entry.team),
                normalize_text(entry.name),
                normalize_text(entry.position),
                float(entry.hours),
            )
        ].append(entry)
    rows: List[dict] = []
    summary: List[dict] = []
    for (entry_date, _, _, _, _), items in grouped.items():
        representative = items[0]
        all_sources = sorted(date_sources.get(entry_date, set()))
        sources_present = sorted({item.source_label for item in items})
        missing = sorted(set(all_sources) - set(sources_present))
        status, unique_flag = describe_presence(sources_present, missing, len(all_sources))
        for item in items:
            rows.append(
                {
                    "diary_date": entry_date.isoformat(),
                    "team": item.team,
                    "name": item.name,
                    "position": item.position,
                    "hours": item.hours,
                    "source_file": item.source_file,
                    "worksheet": item.worksheet,
                    "unique_to_source": unique_flag,
                    "status": status,
                    "all_sources_for_date": "; ".join(all_sources),
                }
            )
        summary.append(
            {
                "diary_date": entry_date.isoformat(),
                "team": representative.team,
                "name": representative.name,
                "position": representative.position,
                "hours": representative.hours,
                "sources_present": "; ".join(sources_present),
                "sources_missing": "; ".join(missing),
                "status": status,
                "report_copies_for_date": len(all_sources),
            }
        )
    rows.sort(key=lambda row: (row["diary_date"], row["team"], row["name"], row["source_file"]))
    summary.sort(key=lambda row: (row["diary_date"], row["team"], row["name"]))
    return rows, summary


def describe_presence(
    sources_present: Sequence[str], missing_sources: Sequence[str], total_sources: int
) -> Tuple[str, bool]:
    unique_sources = list(dict.fromkeys(sources_present))
    present_count = len(unique_sources)
    if total_sources <= 1:
        return ("only available copy", False)
    if present_count == total_sources:
        return (f"present in all {total_sources} copies", False)
    if present_count == 1:
        current = unique_sources[0] if unique_sources else "unknown source"
        missing_note = ", ".join(missing_sources)
        suffix = f" - missing from: {missing_note}" if missing_note else ""
        return (f"single instance in {current}{suffix}", True)
    current_sources = ", ".join(unique_sources)
    missing_note = ", ".join(missing_sources)
    suffix = f"; missing from: {missing_note}" if missing_note else ""
    return (f"in {present_count}/{total_sources} copies ({current_sources}){suffix}", False)


def write_csv(path: Path, headers: List[str], rows: Iterable[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def split_multiline(value: str) -> List[str]:
    pieces = []
    for chunk in value.replace("\r", "").split("\n"):
        cleaned = chunk.strip()
        if cleaned:
            pieces.append(cleaned)
    if not pieces and value.strip():
        pieces.append(value.strip())
    return pieces


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


if __name__ == "__main__":
    main()
