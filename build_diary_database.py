#!/usr/bin/env python3
"""
Build a SQLite database containing activities, personnel, and delay/issue entries
from the project diary Excel workbooks.

Example:
    python build_diary_database.py --root "." --database diary.sqlite --reset
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook


DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d/%m/%y",
    "%d-%m-%Y",
    "%d-%m-%y",
    "%d.%m.%Y",
    "%d.%m.%y",
    "%d %B %Y",
    "%d %b %Y",
    "%B %d %Y",
    "%b %d %Y",
]

NUMERIC_DATE_RE = re.compile(r"(?<!\d)(\d{1,2})[\\/.\-](\d{1,2})[\\/.\-](\d{2,4})(?!\d)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create SQLite database from diary Excel files.")
    parser.add_argument("--root", default=".", help="Root folder that contains the project data.")
    parser.add_argument(
        "--client-dir",
        default="001-Client reports",
        help="Relative path (under root) to the client report workbooks.",
    )
    parser.add_argument(
        "--supervisor-dir",
        default="002-Supervisor_Reports",
        help="Relative path (under root) to the supervisor report workbooks.",
    )
    parser.add_argument(
        "--database",
        default="diary.sqlite",
        help="Path to the SQLite database that will be created/populated.",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Delete rows for the affected diary dates before ingesting.",
    )
    parser.add_argument(
        "--use-supervisor",
        action="store_true",
        help="Enable supervisor ingestion (column K comments + extension notes).",
    )
    parser.add_argument(
        "--use-client-fallback",
        action="store_true",
        help=(
            "Enable storing client PRODUCTION entries as fallback when no supervisor report exists for a date."
        ),
    )
    parser.add_argument(
        "--skip-supervisor",
        action="store_true",
        help="Convenience flag to disable supervisor ingestion (overrides --use-supervisor).",
    )
    parser.add_argument(
        "--skip-client-fallback",
        action="store_true",
        help="Convenience flag to disable fallback ingestion (overrides --use-client-fallback).",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Run ingestion validation against the existing database without modifying any data.",
    )
    return parser.parse_args()


@dataclass
class SheetRow:
    raw: Tuple[Optional[object], ...]
    text: Tuple[str, ...]
    joined: str
    upper: str


@dataclass
class ClientSheetData:
    diary_date: date
    source_file: str
    worksheet: str
    activities: List[str]
    personnel: List[Tuple[str, str, str, float]]
    delays: List[Tuple[str, str, float, str]]
    incidents: List[Tuple[str, str, float, str]]


@dataclass
class SupervisorCommentRecord:
    diary_date: date
    label: str
    hours: Optional[float]
    machine: str
    start_smu: str
    end_smu: str
    machine_hours: str
    location: str
    activity: str
    material: str
    comment: str
    source_file: str
    worksheet: str


@dataclass
class SupervisorSheetData:
    diary_date: date
    source_file: str
    worksheet: str
    comments: List[SupervisorCommentRecord]
    extension_notes: List[str]


@dataclass
class FallbackActivity:
    diary_date: date
    text: str
    source_file: str
    worksheet: str


class DiaryDatabase:
    def __init__(self, path: Path, reset: bool = False) -> None:
        if reset and path.exists():
            path.unlink()
        self.conn = sqlite3.connect(path)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._create_schema()

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS activities (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                activity TEXT NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                UNIQUE(diary_date, activity)
            );

            CREATE TABLE IF NOT EXISTS personnel (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                team_type TEXT NOT NULL,
                name TEXT NOT NULL,
                position TEXT NOT NULL,
                hours REAL NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                UNIQUE(diary_date, team_type, name)
            );

            CREATE TABLE IF NOT EXISTS delays_issues (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                entry_type TEXT NOT NULL CHECK(entry_type IN ('delay', 'issue')),
                label TEXT NOT NULL,
                description TEXT NOT NULL,
                qty REAL NOT NULL,
                comments TEXT NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                UNIQUE(diary_date, entry_type, label, description, qty, comments)
            );

            CREATE TABLE IF NOT EXISTS supervisor_comments (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                worker_or_group TEXT NOT NULL,
                hours REAL,
                machine TEXT NOT NULL,
                start_smu TEXT NOT NULL,
                end_smu TEXT NOT NULL,
                machine_hours TEXT NOT NULL,
                location TEXT NOT NULL,
                activity TEXT NOT NULL,
                material TEXT NOT NULL,
                comment TEXT NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                audit_status TEXT,
                audit_model TEXT,
                audit_timestamp TEXT,
                audit_notes TEXT,
                UNIQUE(diary_date, worker_or_group, comment, source_file, worksheet)
            );

            CREATE TABLE IF NOT EXISTS supervisor_extension_notes (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                note TEXT NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                UNIQUE(diary_date, note, source_file, worksheet)
            );

            CREATE TABLE IF NOT EXISTS client_fallback_activities (
                id INTEGER PRIMARY KEY,
                diary_date TEXT NOT NULL,
                activity TEXT NOT NULL,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL,
                UNIQUE(diary_date, activity, source_file, worksheet)
            );
            """
        )
        self._ensure_audit_columns()

    def _ensure_audit_columns(self) -> None:
        self._ensure_column("supervisor_comments", "audit_status", "TEXT")
        self._ensure_column("supervisor_comments", "audit_model", "TEXT")
        self._ensure_column("supervisor_comments", "audit_timestamp", "TEXT")
        self._ensure_column("supervisor_comments", "audit_notes", "TEXT")

    def _ensure_column(self, table: str, column: str, definition: str) -> None:
        cur = self.conn.execute(f"PRAGMA table_info({table})")
        columns = {row[1] for row in cur.fetchall()}
        if column not in columns:
            self.conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    def insert_activity(self, diary_date: str, activity: str, source_file: str, worksheet: str) -> bool:
        if not activity:
            return False
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO activities (diary_date, activity, source_file, worksheet)
            VALUES (?, ?, ?, ?)
            """,
            (diary_date, activity, source_file, worksheet),
        )
        return cur.rowcount > 0

    def insert_person(self, diary_date: str, team: str, name: str, position: str, hours: float, source_file: str, worksheet: str) -> bool:
        if not name:
            return False
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO personnel (diary_date, team_type, name, position, hours, source_file, worksheet)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (diary_date, team, name.strip(), position or "", float(hours or 0), source_file, worksheet),
        )
        return cur.rowcount > 0

    def insert_delay_issue(
        self,
        diary_date: str,
        entry_type: str,
        label: str,
        description: str,
        qty: float,
        comments: str,
        source_file: str,
        worksheet: str,
    ) -> bool:
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO delays_issues
                (diary_date, entry_type, label, description, qty, comments, source_file, worksheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (diary_date, entry_type, label, description, float(qty or 0), comments, source_file, worksheet),
        )
        return cur.rowcount > 0

    def insert_supervisor_comment(self, record: SupervisorCommentRecord) -> bool:
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO supervisor_comments
                (diary_date, worker_or_group, hours, machine, start_smu, end_smu, machine_hours, location, activity, material, comment, source_file, worksheet)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record.diary_date.isoformat(),
                record.label,
                record.hours,
                record.machine,
                record.start_smu,
                record.end_smu,
                record.machine_hours,
                record.location,
                record.activity,
                record.material,
                record.comment,
                record.source_file,
                record.worksheet,
            ),
        )
        return cur.rowcount > 0

    def insert_extension_note(self, record: SupervisorSheetData, note: str) -> bool:
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO supervisor_extension_notes
                (diary_date, note, source_file, worksheet)
            VALUES (?, ?, ?, ?)
            """,
            (record.diary_date.isoformat(), note, record.source_file, record.worksheet),
        )
        return cur.rowcount > 0

    def insert_fallback_activity(self, entry: FallbackActivity) -> bool:
        cur = self.conn.execute(
            """
            INSERT OR IGNORE INTO client_fallback_activities
                (diary_date, activity, source_file, worksheet)
            VALUES (?, ?, ?, ?)
            """,
            (entry.diary_date.isoformat(), entry.text, entry.source_file, entry.worksheet),
        )
        return cur.rowcount > 0

    def delete_dates(self, diary_dates: Set[str]) -> None:
        if not diary_dates:
            return
        placeholders = ",".join("?" for _ in diary_dates)
        for table in [
            "activities",
            "personnel",
            "delays_issues",
            "supervisor_comments",
            "supervisor_extension_notes",
            "client_fallback_activities",
        ]:
            self.conn.execute(
                f"DELETE FROM {table} WHERE diary_date IN ({placeholders})",
                tuple(sorted(diary_dates)),
            )

    def commit(self) -> None:
        self.conn.commit()


def iter_excel_files(directory: Path) -> List[Path]:
    files: List[Path] = []
    for file_path in sorted(directory.rglob("*.xlsx")):
        if file_path.name.startswith("~$"):
            continue
        files.append(file_path)
    return files


def iter_sheet_rows(ws) -> Iterable[SheetRow]:
    for row in ws.iter_rows(values_only=True):
        raw = tuple(row)
        text = tuple(_format_cell(value) for value in raw)
        if not any(text):
            continue
        joined = " ".join(value.strip() for value in text if value)
        if not joined:
            continue
        yield SheetRow(raw=raw, text=text, joined=joined, upper=joined.upper())


def _format_cell(value: Optional[object]) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _row_text(worksheet, row_index: int) -> str:
    values: List[str] = []
    for cell in worksheet[row_index]:
        text = _format_cell(cell.value)
        if text:
            values.append(text)
    return " | ".join(values)


def _text(value: Optional[object]) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def extract_diary_date(rows: Sequence[SheetRow]) -> Optional[date]:
    for row in rows:
        for value in row.raw:
            parsed = _parse_date_value(value)
            if parsed:
                return parsed
    blob = " ".join(row.joined for row in rows)
    return _parse_date_from_string(blob)


def _parse_date_value(value: Optional[object]) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _parse_date_from_string(value)
    return None


def _parse_date_from_string(value: str) -> Optional[date]:
    cleaned = value.strip()
    if not cleaned:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", cleaned):
        try:
            return datetime.fromisoformat(cleaned).date()
        except ValueError:
            pass
    match = NUMERIC_DATE_RE.search(cleaned)
    if match:
        day, month, year = match.groups()
        year_value = int(year)
        if year_value < 100:
            year_value += 2000
        return date(year_value, int(month), int(day))
    return None


def extract_activities(rows: Sequence[SheetRow]) -> List[str]:
    activities: List[str] = []
    in_section = False
    for row in rows:
        if not in_section:
            if "PRODUCTION" in row.upper:
                in_section = True
            continue
        if "PHOTOS" in row.upper:
            break
        text = row.joined.strip()
        if not text or "COMMUNICATIONS" in row.upper or "PRODUCTION" in row.upper:
            continue
        activities.append(text)
    return activities


def extract_personnel(rows: Sequence[SheetRow]) -> List[Tuple[str, str, str, float]]:
    people: List[Tuple[str, str, str, float]] = []
    start_idx = next((idx for idx, row in enumerate(rows) if "PERSONNEL" in row.upper), None)
    if start_idx is None:
        return people
    group_row: Optional[SheetRow] = None
    header_seen = False
    for row in rows[start_idx + 1 :]:
        if "PLANT" in row.upper:
            break
        if group_row is None:
            group_row = row
            continue
        if not header_seen:
            lower_values = [val.lower() for val in row.text if val]
            if "name" in lower_values and "position" in lower_values:
                header_seen = True
                continue
        if not any(row.text):
            continue
        if any(cell.lower().startswith("total") for cell in row.text if cell):
            continue
        for col, team in _group_columns(group_row):
            name = _safe_get(row.raw, col)
            if not _is_valid_name(name):
                continue
            position = _safe_get(row.raw, col + 1) or ""
            hours = _to_number(_safe_get(row.raw, col + 2))
            people.append((team, str(name).strip(), str(position).strip(), hours if hours is not None else 0.0))
    return people


def _group_columns(group_row: SheetRow) -> List[Tuple[int, str]]:
    groups: List[Tuple[int, str]] = []
    for idx, cell in enumerate(group_row.text):
        if cell:
            groups.append((idx, cell))
    collapsed: List[Tuple[int, str]] = []
    for idx, label in groups:
        if not collapsed or idx - collapsed[-1][0] >= 3:
            collapsed.append((idx, label))
    return collapsed


def _safe_get(row: Tuple[Optional[object], ...], index: int) -> Optional[object]:
    if index < len(row):
        return row[index]
    return None


def _is_valid_name(value: Optional[object]) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text or text.isdigit():
        return False
    return text.lower() not in {"name", "contact"}


def _to_number(value: Optional[object]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def extract_delay_issue_rows(rows: Sequence[SheetRow]) -> List[Tuple[str, str, float, str]]:
    entries: List[Tuple[str, str, float, str]] = []
    in_delays = False
    for row in rows:
        if not in_delays:
            if "DELAYS-OPPORTUNITY" in row.upper:
                in_delays = True
            continue
        if "HSEQ" in row.upper:
            break
        text = row.joined.strip()
        if not text:
            continue
        for chunk in _split_multiline(text):
            entries.append(("delay", chunk, 0.0, ""))
    return entries


def _split_multiline(value: str) -> List[str]:
    parts = []
    for piece in value.replace("\r", "").split("\n"):
        chunk = piece.strip()
        if chunk:
            parts.append(chunk)
    if not parts and value.strip():
        parts.append(value.strip())
    return parts


def _should_stop_labour_section(label_upper: str) -> bool:
    if not label_upper:
        return False
    stop_markers = [
        "PLANT NOT",
        "PLANNED WORKS",
        "INCIDENTS",
        "COMMUNICATIONS",
        "DAILY WORK EXTENSION",
        "DAILY WORK PHOTOS",
    ]
    return any(marker in label_upper for marker in stop_markers)


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower()) if value else ""


def extract_incidents(rows: Sequence[SheetRow]) -> List[Tuple[str, str, float, str]]:
    entries: List[Tuple[str, str, float, str]] = []
    in_section = False
    header_skipped = False
    for row in rows:
        if not in_section:
            if "INCIDENTS" in row.upper:
                in_section = True
            continue
        if not header_skipped:
            if "QTY" in row.upper or "COMMENTS" in row.upper:
                header_skipped = True
                continue
        if "COMMUNICATIONS" in row.upper or "PRODUCTION" in row.upper:
            break
        label = row.text[0] if row.text else ""
        if not label:
            continue
        qty = _to_number(_safe_get(row.raw, 1)) or 0.0
        comments = row.text[2] if len(row.text) > 2 else ""
        comments_clean = comments.strip()
        if comments_clean.upper() in {"NA", "N/A"}:
            comments_clean = ""
        if qty == 0.0 and not comments_clean:
            continue
        entries.append(("issue", label.strip(), qty, comments_clean))
    return entries


def parse_client_sheets(client_root: Path) -> List[ClientSheetData]:
    entries: List[ClientSheetData] = []
    if not client_root.exists():
        return entries
    for file_path in iter_excel_files(client_root):
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Failed to open {file_path}: {exc}")
            continue
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = extract_diary_date(rows)
            if diary_date is None:
                continue
            entries.append(
                ClientSheetData(
                    diary_date=diary_date,
                    source_file=str(file_path),
                    worksheet=sheet_name,
                    activities=extract_activities(rows),
                    personnel=extract_personnel(rows),
                    delays=extract_delay_issue_rows(rows),
                    incidents=extract_incidents(rows),
                )
            )
    return entries


def ingest_client(db: DiaryDatabase, sheets: Iterable[ClientSheetData], stats: Dict[str, object]) -> None:
    grouped: Dict[date, List[ClientSheetData]] = defaultdict(list)
    for sheet in sheets:
        grouped[sheet.diary_date].append(sheet)
    for diary_date, date_sheets in grouped.items():
        diary_date_str = diary_date.isoformat()
        activities_seen: Set[str] = set()
        personnel_seen: Set[Tuple[str, str, str, float]] = set()
        delays_seen: Set[Tuple[str, str, float, str]] = set()
        incidents_seen: Set[Tuple[str, str, float, str]] = set()
        for sheet in sorted(date_sheets, key=lambda item: (item.source_file, item.worksheet)):
            for activity in sheet.activities:
                norm = _normalize_text(activity)
                if not norm or norm in activities_seen:
                    continue
                activities_seen.add(norm)
                if db.insert_activity(diary_date_str, activity, sheet.source_file, sheet.worksheet):
                    stats["activities"] += 1
            for team, name, position, hours in sheet.personnel:
                key = (
                    _normalize_text(team),
                    _normalize_text(name),
                    _normalize_text(position),
                    float(hours or 0),
                )
                if key in personnel_seen:
                    continue
                personnel_seen.add(key)
                if db.insert_person(diary_date_str, team, name, position, hours, sheet.source_file, sheet.worksheet):
                    stats["personnel"] += 1
            for entry_type, label, qty, comments in sheet.delays:
                key = (entry_type, _normalize_text(label or ""), float(qty or 0), _normalize_text(comments or ""))
                if key in delays_seen:
                    continue
                delays_seen.add(key)
                if db.insert_delay_issue(diary_date_str, entry_type, "", label, qty, comments, sheet.source_file, sheet.worksheet):
                    stats["delays_issues"] += 1
            for entry_type, label, qty, comments in sheet.incidents:
                key = (entry_type, _normalize_text(label or ""), float(qty or 0), _normalize_text(comments or ""))
                if key in incidents_seen:
                    continue
                incidents_seen.add(key)
                if db.insert_delay_issue(diary_date_str, entry_type, label, "", qty, comments, sheet.source_file, sheet.worksheet):
                    stats["delays_issues"] += 1


def parse_supervisor_reports(supervisor_root: Path) -> List[SupervisorSheetData]:
    sheets: List[SupervisorSheetData] = []
    if not supervisor_root.exists():
        return sheets
    for file_path in iter_excel_files(supervisor_root):
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Failed to open {file_path}: {exc}")
            continue
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = extract_diary_date(rows)
            if diary_date is None:
                continue
            comments = extract_supervisor_comments(worksheet, diary_date, str(file_path), sheet_name)
            extension = extract_extension_notes(worksheet)
            sheets.append(
                SupervisorSheetData(
                    diary_date=diary_date,
                    source_file=str(file_path),
                    worksheet=sheet_name,
                    comments=comments,
                    extension_notes=extension,
                )
            )
    return sheets


def ingest_supervisor(db: DiaryDatabase, sheets: Iterable[SupervisorSheetData], stats: Dict[str, object]) -> None:
    for sheet in sheets:
        for record in sheet.comments:
            if db.insert_supervisor_comment(record):
                stats["supervisor_comments"] += 1
        for note in sheet.extension_notes:
            if db.insert_extension_note(sheet, note):
                stats["supervisor_extension_notes"] += 1


def parse_client_fallback(client_root: Path, skip_dates: Set[date]) -> List[FallbackActivity]:
    entries: List[FallbackActivity] = []
    if not client_root.exists():
        return entries
    grouped: Dict[date, List[FallbackActivity]] = defaultdict(list)
    seen: Dict[date, Set[str]] = defaultdict(set)
    for file_path in iter_excel_files(client_root):
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Failed to open {file_path}: {exc}")
            continue
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = extract_diary_date(rows)
            if diary_date is None or diary_date in skip_dates:
                continue
            for text in extract_activities(rows):
                norm = _normalize_text(text)
                if not norm or norm in seen[diary_date]:
                    continue
                seen[diary_date].add(norm)
                grouped[diary_date].append(
                    FallbackActivity(
                        diary_date=diary_date,
                        text=text,
                        source_file=str(file_path),
                        worksheet=sheet_name,
                    )
                )
    for date_key in sorted(grouped):
        entries.extend(grouped[date_key])
    return entries


def ingest_fallback(db: DiaryDatabase, entries: Iterable[FallbackActivity], stats: Dict[str, object]) -> None:
    for entry in entries:
        if db.insert_fallback_activity(entry):
            stats["fallback_activities"] += 1


def validate_ingest(db: DiaryDatabase, *, require_coverage: bool = True) -> None:
    issues: List[str] = []
    cur = db.conn.cursor()
    tables = [
        "activities",
        "personnel",
        "delays_issues",
        "supervisor_comments",
        "supervisor_extension_notes",
        "client_fallback_activities",
    ]
    for table in tables:
        cur.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE name = ?", (table,))
        exists = cur.fetchone()[0]
        if not exists:
            continue
        cur.execute(f"SELECT COUNT(*) FROM {table} WHERE diary_date IS NULL OR diary_date = ''")
        nulls = cur.fetchone()[0]
        if nulls:
            issues.append(f"{table} has {nulls} rows with missing diary_date")
    cur.execute(
        """
        SELECT COUNT(*)
        FROM client_fallback_activities
        WHERE diary_date IN (SELECT DISTINCT diary_date FROM supervisor_comments)
        """
    )
    conflicts = cur.fetchone()[0]
    if conflicts:
        issues.append("Fallback activities include dates that also have supervisor reports")
    if require_coverage:
        cur.execute(
            """
            WITH source_dates AS (
                SELECT diary_date FROM activities
                UNION SELECT diary_date FROM personnel
                UNION SELECT diary_date FROM delays_issues
            )
            SELECT diary_date
            FROM source_dates
            WHERE diary_date NOT IN (
                SELECT diary_date FROM supervisor_comments
                UNION SELECT diary_date FROM supervisor_extension_notes
                UNION SELECT diary_date FROM client_fallback_activities
            )
            """
        )
        uncovered = [row[0] for row in cur.fetchall()]
        if uncovered:
            issues.append(
                "Missing supervisor and fallback coverage on dates: "
                + ", ".join(sorted(uncovered)[:10])
                + ("..." if len(uncovered) > 10 else "")
            )
    cur.execute(
        """
        WITH fallback_counts AS (
            SELECT diary_date, COUNT(*) AS cnt FROM client_fallback_activities GROUP BY diary_date
        ),
        activity_counts AS (
            SELECT diary_date, COUNT(*) AS cnt FROM activities GROUP BY diary_date
        )
        SELECT f.diary_date, f.cnt, a.cnt
        FROM fallback_counts f
        JOIN activity_counts a ON a.diary_date = f.diary_date
        WHERE f.cnt > a.cnt
        """
    )
    excessive = cur.fetchall()
    if excessive:
        issues.append(
            "Fallback counts exceed activity counts on dates: "
            + ", ".join(f"{row[0]} (fallback {row[1]} > activities {row[2]})" for row in excessive[:5])
            + ("..." if len(excessive) > 5 else "")
        )
    if issues:
        raise RuntimeError("Ingestion validation failed:\n" + "\n".join(f"- {msg}" for msg in issues))


def extract_supervisor_comments(
    worksheet, diary_date: date, source_file: str, sheet_name: str
) -> List[SupervisorCommentRecord]:
    entries: List[SupervisorCommentRecord] = []
    header_row_seen = False
    for row_index in range(1, worksheet.max_row + 1):
        hours_header = _text(worksheet.cell(row_index, 3).value)
        machine_header = _text(worksheet.cell(row_index, 4).value)
        if not header_row_seen:
            if hours_header.lower() == "hours" and machine_header.lower() == "machine":
                header_row_seen = True
            continue
        label = _text(worksheet.cell(row_index, 2).value)
        upper_label = label.upper()
        if _should_stop_labour_section(upper_label):
            break
        comment = _text(worksheet.cell(row_index, 11).value)
        if not comment:
            continue
        entries.append(
            SupervisorCommentRecord(
                diary_date=diary_date,
                label=label,
                hours=_to_number(worksheet.cell(row_index, 3).value),
                machine=_text(worksheet.cell(row_index, 4).value),
                start_smu=_text(worksheet.cell(row_index, 5).value),
                end_smu=_text(worksheet.cell(row_index, 6).value),
                machine_hours=_text(worksheet.cell(row_index, 7).value),
                location=_text(worksheet.cell(row_index, 8).value),
                activity=_text(worksheet.cell(row_index, 9).value),
                material=_text(worksheet.cell(row_index, 10).value),
                comment=comment,
                source_file=source_file,
                worksheet=sheet_name,
            )
        )
    return entries


def extract_extension_notes(worksheet) -> List[str]:
    start_row = None
    end_row = None
    for row_index in range(1, worksheet.max_row + 1):
        row_text = _row_text(worksheet, row_index).lower()
        if start_row is None and "daily work extension" in row_text:
            start_row = row_index + 1
            continue
        if start_row is not None and "daily work photos" in row_text:
            end_row = row_index
            break
    if start_row is None or end_row is None or start_row >= end_row:
        return []
    notes: List[str] = []
    for row_index in range(start_row, end_row):
        note_text = _row_text(worksheet, row_index)
        if note_text:
            notes.append(note_text)
    return notes


def run_ingest(args: argparse.Namespace) -> Dict[str, object]:
    root = Path(args.root).expanduser().resolve()
    client_root = (root / args.client_dir).resolve()
    supervisor_root = (root / args.supervisor_dir).resolve()
    db_path = Path(args.database).expanduser().resolve()

    use_supervisor = args.use_supervisor and not args.skip_supervisor
    use_fallback = args.use_client_fallback and not args.skip_client_fallback

    db = DiaryDatabase(db_path, reset=False)
    stats: Dict[str, object] = {
        "activities": 0,
        "personnel": 0,
        "delays_issues": 0,
        "supervisor_comments": 0,
        "supervisor_extension_notes": 0,
        "fallback_activities": 0,
    }

    client_sheets = parse_client_sheets(client_root)
    supervisor_sheets: List[SupervisorSheetData] = []
    if use_supervisor and supervisor_root.exists():
        supervisor_sheets = parse_supervisor_reports(supervisor_root)
    supervisor_dates = {sheet.diary_date for sheet in supervisor_sheets}
    supervisor_covered_dates = {
        sheet.diary_date for sheet in supervisor_sheets if sheet.comments or sheet.extension_notes
    }

    fallback_entries: List[FallbackActivity] = []
    if use_fallback:
        skip_dates = supervisor_covered_dates if use_supervisor else set()
        fallback_entries = parse_client_fallback(client_root, skip_dates)

    touched_dates: Set[str] = {sheet.diary_date.isoformat() for sheet in client_sheets}
    touched_dates.update(date.isoformat() for date in supervisor_dates)
    touched_dates.update(entry.diary_date.isoformat() for entry in fallback_entries)

    if args.reset:
        db.delete_dates(touched_dates)

    ingest_client(db, client_sheets, stats)
    if use_supervisor:
        ingest_supervisor(db, supervisor_sheets, stats)
    if use_fallback:
        ingest_fallback(db, fallback_entries, stats)

    db.commit()
    validate_ingest(db, require_coverage=use_supervisor or use_fallback)
    stats["database_path"] = str(db_path)
    return stats


def run_validate(database: str, *, require_coverage: bool = True) -> None:
    db_path = Path(database).expanduser().resolve()
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}")
    db = DiaryDatabase(db_path, reset=False)
    validate_ingest(db, require_coverage=require_coverage)


def main() -> None:
    args = parse_args()
    if args.validate_only:
        run_validate(args.database, require_coverage=args.use_supervisor or args.use_client_fallback)
        print("Validation successful.")
        return
    stats = run_ingest(args)
    print("Ingestion complete.")
    print(f"  Activities inserted:          {stats['activities']}")
    print(f"  Personnel inserted:           {stats['personnel']}")
    print(f"  Delay/issue entries inserted: {stats['delays_issues']}")
    print(f"  Supervisor comments:          {stats['supervisor_comments']}")
    print(f"  Supervisor extension notes:   {stats['supervisor_extension_notes']}")
    print(f"  Fallback activities:          {stats['fallback_activities']}")
    print(f"Database path: {stats['database_path']}")


if __name__ == "__main__":
    main()
