#!/usr/bin/env python3
"""
Parse both supervisor and client daily reports, capturing:
  * Column-K work descriptions from supervisor labour tables.
  * Notes between “Daily Work Extension” and “Daily Work Photos”.
  * Client “PRODUCTION” entries, but only when no supervisor report exists for that date.

Outputs CSV summaries under the requested output directory.
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set

from openpyxl import load_workbook

import build_diary_database as diary


@dataclass
class SupervisorComment:
    diary_date: date
    label: str
    hours: Optional[float]
    machine: str
    start_smu: str
    end_smu: str
    machine_hours: str
    location: str
    activity: str
    material: str
    comment: str
    source_file: str
    worksheet: str


@dataclass
class ExtensionNote:
    diary_date: date
    note: str
    source_file: str
    worksheet: str


@dataclass
class ClientActivity:
    diary_date: date
    text: str
    source_file: str
    worksheet: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract structured data from supervisor and client reports.")
    parser.add_argument("--root", default=".", help="Project root that contains the 001/002 folders.")
    parser.add_argument("--output-dir", default="analysis", help="Directory for CSV outputs.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    supervisor_dir = root / "002-Supervisor_Reports"
    client_dir = root / "001-Client reports"

    supervisor_comments, extension_notes, supervisor_dates = parse_supervisor_reports(supervisor_dir, root)
    client_fallback = parse_client_reports(client_dir, root, supervisor_dates)

    write_csv(
        output_dir / "supervisor_comments.csv",
        [
            "diary_date",
            "worker_or_group",
            "hours",
            "machine",
            "start_smu",
            "end_smu",
            "machine_hours",
            "location",
            "activity",
            "material",
            "comment",
            "source_file",
            "worksheet",
        ],
        (
            {
                "diary_date": entry.diary_date.isoformat(),
                "worker_or_group": entry.label,
                "hours": entry.hours if entry.hours is not None else "",
                "machine": entry.machine,
                "start_smu": entry.start_smu,
                "end_smu": entry.end_smu,
                "machine_hours": entry.machine_hours,
                "location": entry.location,
                "activity": entry.activity,
                "material": entry.material,
                "comment": entry.comment,
                "source_file": entry.source_file,
                "worksheet": entry.worksheet,
            }
            for entry in supervisor_comments
        ),
    )

    write_csv(
        output_dir / "supervisor_daily_extension.csv",
        ["diary_date", "note", "source_file", "worksheet"],
        (
            {
                "diary_date": entry.diary_date.isoformat(),
                "note": entry.note,
                "source_file": entry.source_file,
                "worksheet": entry.worksheet,
            }
            for entry in extension_notes
        ),
    )

    write_csv(
        output_dir / "client_fallback_production.csv",
        ["diary_date", "activity_text", "source_file", "worksheet"],
        (
            {
                "diary_date": entry.diary_date.isoformat(),
                "activity_text": entry.text,
                "source_file": entry.source_file,
                "worksheet": entry.worksheet,
            }
            for entry in client_fallback
        ),
    )

    print("Supervisor reports parsed:")
    print(f"  Labour comments: {len(supervisor_comments)} rows")
    print(f"  Extension notes: {len(extension_notes)} rows")
    print(f"  Unique supervisor dates: {len(supervisor_dates)}")
    print("Client reports parsed (fallback only):")
    print(f"  Production entries: {len(client_fallback)} rows")
    print(f"CSV output directory: {output_dir}")


def parse_supervisor_reports(
    supervisor_dir: Path, root: Path
) -> tuple[list[SupervisorComment], list[ExtensionNote], Set[date]]:
    comments: List[SupervisorComment] = []
    extension_notes: List[ExtensionNote] = []
    dates: Set[date] = set()

    for workbook_path in sorted(supervisor_dir.rglob("*.xlsx")):
        if workbook_path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Failed to open {workbook_path}: {exc}")
            continue
        relative = str(_safe_relative(workbook_path, root))
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(diary.iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = diary.extract_diary_date(rows)
            if diary_date is None:
                continue
            dates.add(diary_date)
            comments.extend(extract_supervisor_comments(worksheet, diary_date, relative, sheet_name))
            extension_notes.extend(extract_extension_notes(worksheet, diary_date, relative, sheet_name))
    comments.sort(key=lambda entry: (entry.diary_date, entry.label))
    extension_notes.sort(key=lambda entry: (entry.diary_date, entry.note))
    return comments, extension_notes, dates


def extract_supervisor_comments(
    worksheet, diary_date: date, source_file: str, sheet_name: str
) -> List[SupervisorComment]:
    entries: List[SupervisorComment] = []
    header_row_seen = False
    for row_index in range(1, worksheet.max_row + 1):
        if not header_row_seen:
            hours_header = _text(worksheet.cell(row_index, 3).value)
            machine_header = _text(worksheet.cell(row_index, 4).value)
            if hours_header.lower() == "hours" and machine_header.lower() == "machine":
                header_row_seen = True
            continue
        label = _text(worksheet.cell(row_index, 2).value)
        upper_label = label.upper()
        if _should_stop_labour_section(upper_label):
            break
        comment = _text(worksheet.cell(row_index, 11).value)
        if not comment:
            continue
        entry = SupervisorComment(
            diary_date=diary_date,
            label=label,
            hours=_to_float(worksheet.cell(row_index, 3).value),
            machine=_text(worksheet.cell(row_index, 4).value),
            start_smu=_text(worksheet.cell(row_index, 5).value),
            end_smu=_text(worksheet.cell(row_index, 6).value),
            machine_hours=_text(worksheet.cell(row_index, 7).value),
            location=_text(worksheet.cell(row_index, 8).value),
            activity=_text(worksheet.cell(row_index, 9).value),
            material=_text(worksheet.cell(row_index, 10).value),
            comment=comment,
            source_file=source_file,
            worksheet=sheet_name,
        )
        entries.append(entry)
    return entries


def extract_extension_notes(
    worksheet, diary_date: date, source_file: str, sheet_name: str
) -> List[ExtensionNote]:
    start_row = None
    end_row = None
    for row_index in range(1, worksheet.max_row + 1):
        row_text = _row_text(worksheet, row_index).lower()
        if start_row is None and "daily work extension" in row_text:
            start_row = row_index + 1
            continue
        if start_row is not None and "daily work photos" in row_text:
            end_row = row_index
            break
    if start_row is None or end_row is None or start_row >= end_row:
        return []
    notes: List[ExtensionNote] = []
    for row_index in range(start_row, end_row):
        note_text = _row_text(worksheet, row_index)
        if not note_text:
            continue
        notes.append(
            ExtensionNote(
                diary_date=diary_date,
                note=note_text,
                source_file=source_file,
                worksheet=sheet_name,
            )
        )
    return notes


def parse_client_reports(
    client_dir: Path, root: Path, supervisor_dates: Set[date]
) -> List[ClientActivity]:
    activities: List[ClientActivity] = []
    for workbook_path in sorted(client_dir.rglob("*.xlsx")):
        if workbook_path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Failed to open {workbook_path}: {exc}")
            continue
        relative = str(_safe_relative(workbook_path, root))
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(diary.iter_sheet_rows(worksheet))
            if not rows:
                continue
            diary_date = diary.extract_diary_date(rows)
            if diary_date is None or diary_date in supervisor_dates:
                continue
            for text in diary.extract_activities(rows):
                activities.append(
                    ClientActivity(
                        diary_date=diary_date,
                        text=text,
                        source_file=relative,
                        worksheet=sheet_name,
                    )
                )
    activities.sort(key=lambda entry: (entry.diary_date, entry.text))
    return activities


def write_csv(path: Path, headers: Sequence[str], rows: Iterable[Dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _text(value: Optional[object]) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _row_text(worksheet, row_index: int) -> str:
    parts: List[str] = []
    for cell in worksheet[row_index]:
        text = _text(cell.value)
        if text:
            parts.append(text)
    return " | ".join(parts)


def _to_float(value: Optional[object]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def _should_stop_labour_section(label_upper: str) -> bool:
    if not label_upper:
        return False
    stop_markers = [
        "PLANT NOT",
        "PLANNED WORKS",
        "INCIDENTS",
        "COMMUNICATIONS",
        "DAILY WORK EXTENSION",
        "DAILY WORK PHOTOS",
    ]
    return any(marker in label_upper for marker in stop_markers)


def _safe_relative(path: Path, root: Path) -> Path:
    try:
        return path.relative_to(root)
    except ValueError:
        return path


if __name__ == "__main__":
    main()
